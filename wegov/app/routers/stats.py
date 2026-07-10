import io
from datetime import date, timedelta
from collections import defaultdict

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.core.database import get_db
from app.core.deps import require_leader_or_above
from app.models.work_record import WorkRecord
from app.models.worker import Worker
from app.models.correction import Correction

router = APIRouter(prefix="/stats", tags=["통계"])


# ─────────────────────────────────────────────
# 1. 공정별 집계
# ─────────────────────────────────────────────
@router.get("/by-process", summary="공정별 통계")
async def stats_by_process(
    start_date: date = Query(..., description="시작 날짜"),
    end_date: date = Query(..., description="종료 날짜"),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_leader_or_above),
):
    result = await db.execute(
        select(
            WorkRecord.process,
            func.count(WorkRecord.id).label("count"),
            func.count(WorkRecord.worker_id.distinct()).label("worker_count"),
            func.sum(WorkRecord.duration_minutes).label("total_minutes"),
            func.avg(WorkRecord.duration_minutes).label("avg_minutes"),
        )
        .where(
            and_(
                WorkRecord.checkin_at >= start_date,
                WorkRecord.checkin_at < end_date + timedelta(days=1),
                WorkRecord.is_voided.is_(False),
                WorkRecord.checkout_at.isnot(None),
            )
        )
        .group_by(WorkRecord.process)
        .order_by(WorkRecord.process)
    )
    rows = result.all()
    return [
        {
            "process": r.process,
            "count": r.count,
            "worker_count": r.worker_count,
            "total_minutes": r.total_minutes or 0,
            "avg_minutes": round(r.avg_minutes or 0, 1),
        }
        for r in rows
    ]


# ─────────────────────────────────────────────
# 2. 사업장(고객사)별 집계
# ─────────────────────────────────────────────
@router.get("/by-site", summary="사업장별 통계")
async def stats_by_site(
    start_date: date = Query(..., description="시작 날짜"),
    end_date: date = Query(..., description="종료 날짜"),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_leader_or_above),
):
    result = await db.execute(
        select(
            WorkRecord.site_name,
            func.count(WorkRecord.id).label("count"),
            func.count(WorkRecord.worker_id.distinct()).label("worker_count"),
            func.sum(WorkRecord.duration_minutes).label("total_minutes"),
        )
        .where(
            and_(
                WorkRecord.checkin_at >= start_date,
                WorkRecord.checkin_at < end_date + timedelta(days=1),
                WorkRecord.is_voided.is_(False),
                WorkRecord.checkout_at.isnot(None),
            )
        )
        .group_by(WorkRecord.site_name)
        .order_by(WorkRecord.site_name)
    )
    rows = result.all()
    return [
        {
            "site_name": r.site_name,
            "count": r.count,
            "worker_count": r.worker_count,
            "total_minutes": r.total_minutes or 0,
        }
        for r in rows
    ]


# ─────────────────────────────────────────────
# 3. 작업자별 월간 집계
# ─────────────────────────────────────────────
@router.get("/by-worker", summary="작업자별 월간 통계")
async def stats_by_worker(
    year: int = Query(..., description="연도"),
    month: int = Query(..., description="월"),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_leader_or_above),
):
    start = date(year, month, 1)
    end = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)

    records_result = await db.execute(
        select(
            WorkRecord.worker_id,
            func.sum(WorkRecord.duration_minutes).label("total_minutes"),
            func.count(WorkRecord.id).label("total_count"),
        )
        .where(
            and_(
                WorkRecord.checkin_at >= start,
                WorkRecord.checkin_at < end,
                WorkRecord.is_voided.is_(False),
                WorkRecord.checkout_at.isnot(None),
            )
        )
        .group_by(WorkRecord.worker_id)
    )
    records_rows = records_result.all()
    worker_ids = [r.worker_id for r in records_rows]

    miss_result = await db.execute(
        select(WorkRecord.worker_id, func.count(WorkRecord.id).label("miss_count"))
        .where(
            and_(
                WorkRecord.checkin_at >= start,
                WorkRecord.checkin_at < end,
                WorkRecord.checkout_at.is_(None),
                WorkRecord.is_voided.is_(False),
            )
        )
        .group_by(WorkRecord.worker_id)
    )
    correction_result = await db.execute(
        select(Correction.worker_id, func.count(Correction.id).label("correction_count"))
        .where(
            and_(
                Correction.created_at >= start,
                Correction.created_at < end,
                Correction.status == "approved",
            )
        )
        .group_by(Correction.worker_id)
    )
    miss_map = {r.worker_id: r.miss_count for r in miss_result.all()}
    correction_map = {r.worker_id: r.correction_count for r in correction_result.all()}

    workers_result = await db.execute(
        select(Worker.id, Worker.name, Worker.company).where(Worker.id.in_(worker_ids))
    )
    worker_map = {w.id: {"name": w.name, "company": w.company} for w in workers_result.all()}

    return [
        {
            "worker_id": r.worker_id,
            "name": worker_map.get(r.worker_id, {}).get("name", ""),
            "company": worker_map.get(r.worker_id, {}).get("company", ""),
            "total_minutes": r.total_minutes or 0,
            "total_count": r.total_count,
            "miss_count": miss_map.get(r.worker_id, 0),
            "correction_count": correction_map.get(r.worker_id, 0),
        }
        for r in records_rows
    ]


# ─────────────────────────────────────────────
# 4. 엑셀 다운로드
# ─────────────────────────────────────────────
@router.get("/excel", summary="월간 출입기록 엑셀 다운로드")
async def download_excel(
    year: int = Query(..., description="연도"),
    month: int = Query(..., description="월"),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_leader_or_above),
):
    start = date(year, month, 1)
    end = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)

    result = await db.execute(
        select(WorkRecord, Worker)
        .join(Worker, WorkRecord.worker_id == Worker.id)
        .where(
            and_(
                WorkRecord.checkin_at >= start,
                WorkRecord.checkin_at < end,
                WorkRecord.is_voided.is_(False),
            )
        )
        .order_by(Worker.name, WorkRecord.checkin_at, WorkRecord.round)
    )
    rows = result.all()

    grouped: dict = defaultdict(lambda: defaultdict(list))
    worker_info: dict = {}
    for record, worker in rows:
        d = record.checkin_at.date()
        grouped[worker.id][d].append(record)
        worker_info[worker.id] = {"name": worker.name, "company": worker.company}

    wb = Workbook()

    # ── 시트 1: 직원별 일일 상세기록 ──
    ws1 = wb.active
    ws1.title = f"{year}년{month}월 출입기록"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    headers = [
        "이름", "회사", "날짜", "근무표",
        "1입실", "1퇴실", "1작업시간",
        "2입실", "2퇴실", "2작업시간",
        "3입실", "3퇴실", "3작업시간",
        "4입실", "4퇴실", "4작업시간",
        "5입실", "5퇴실", "5작업시간",
        "6입실", "6퇴실", "6작업시간",
        "휴게시간1", "휴게시간2", "휴게시간3", "휴게시간4", "휴게시간5",
        "총작업시간(분)", "총휴게시간(분)", "실제작업시간(분)", "근무시간(분)",
        "지각여부", "비고",
    ]
    ws1.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for worker_id, dates in sorted(grouped.items(), key=lambda x: worker_info[x[0]]["name"]):
        info = worker_info[worker_id]
        for d in sorted(dates.keys()):
            recs = sorted(dates[d], key=lambda r: r.round)
            row = [info["name"], info["company"], d.strftime("%Y-%m-%d"), recs[0].shift_type if recs else ""]

            total_work = 0
            for i in range(6):
                if i < len(recs):
                    r = recs[i]
                    cin = r.checkin_at.strftime("%H:%M") if r.checkin_at else ""
                    cout = r.checkout_at.strftime("%H:%M") if r.checkout_at else ""
                    dur = r.duration_minutes or 0
                    total_work += dur
                    row += [cin, cout, dur]
                else:
                    row += ["", "", ""]

            rest_times = []
            for i in range(min(len(recs) - 1, 5)):
                if recs[i].checkout_at and recs[i + 1].checkin_at:
                    gap = int((recs[i + 1].checkin_at - recs[i].checkout_at).total_seconds() / 60)
                    rest_times.append(gap)
                else:
                    rest_times.append("")
            while len(rest_times) < 5:
                rest_times.append("")
            row += rest_times

            total_rest = sum(t for t in rest_times if isinstance(t, int))
            actual_work = total_work
            work_time = total_work + total_rest
            row += [total_work, total_rest, actual_work, work_time]

            first_in = recs[0].checkin_at if recs else None
            late = "Y" if first_in and (first_in.hour > 9 or (first_in.hour == 9 and first_in.minute > 0)) else "N"
            row.append(late)

            notes = [r.note for r in recs if r.note]
            row.append(" / ".join(notes) if notes else "")
            ws1.append(row)

    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 14

    # ── 시트 2: 공정별 집계 ──
    ws2 = wb.create_sheet("공정별 집계")
    ws2.append(["공정명", "사업장", "투입 횟수", "투입 인원", "총 작업시간(분)", "평균 작업시간(분)"])
    for col in range(1, 7):
        cell = ws2.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    process_stats: dict = defaultdict(lambda: {"count": 0, "workers": set(), "total": 0})
    for record, worker in rows:
        key = (record.process, record.site_name)
        process_stats[key]["count"] += 1
        process_stats[key]["workers"].add(worker.id)
        process_stats[key]["total"] += record.duration_minutes or 0

    for (process, site), s in sorted(process_stats.items()):
        avg = round(s["total"] / s["count"], 1) if s["count"] else 0
        ws2.append([process, site, s["count"], len(s["workers"]), s["total"], avg])

    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # ── 시트 3: 작업자별 월간 집계 ──
    ws3 = wb.create_sheet("작업자별 집계")
    ws3.append(["이름", "회사", "총 투입 횟수", "총 작업시간(분)", "누락 건수", "수정 승인 건수"])
    for col in range(1, 7):
        cell = ws3.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    miss_result = await db.execute(
        select(WorkRecord.worker_id, func.count(WorkRecord.id).label("cnt"))
        .where(
            and_(
                WorkRecord.checkin_at >= start,
                WorkRecord.checkin_at < end,
                WorkRecord.checkout_at.is_(None),
                WorkRecord.is_voided.is_(False),
            )
        )
        .group_by(WorkRecord.worker_id)
    )
    corr_result = await db.execute(
        select(Correction.worker_id, func.count(Correction.id).label("cnt"))
        .where(
            and_(
                Correction.created_at >= start,
                Correction.created_at < end,
                Correction.status == "approved",
            )
        )
        .group_by(Correction.worker_id)
    )
    miss_map = {r.worker_id: r.cnt for r in miss_result.all()}
    corr_map = {r.worker_id: r.cnt for r in corr_result.all()}

    for worker_id, dates in sorted(grouped.items(), key=lambda x: worker_info[x[0]]["name"]):
        info = worker_info[worker_id]
        all_recs = [r for day_recs in dates.values() for r in day_recs]
        total_minutes = sum(r.duration_minutes or 0 for r in all_recs)
        ws3.append([
            info["name"], info["company"], len(all_recs), total_minutes,
            miss_map.get(worker_id, 0), corr_map.get(worker_id, 0),
        ])

    for col in range(1, 7):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"출입기록_{year}년{month}월.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )
